import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(file_name, position_of_chart, new_file_name):
    workbook = xl.load_workbook(file_name + '.xlsx')
    sheet = workbook['Sheet1']

    for row in range(2, sheet.max_row + 1):
        corrected_price = 0.9 * sheet.cell(row, 3).value
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, position_of_chart)

    workbook.save(new_file_name + '.xlsx')


process_workbook('transactions', 'e2', 'transactions3')
